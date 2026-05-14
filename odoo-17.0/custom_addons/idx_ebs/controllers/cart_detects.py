# -*- coding: utf-8 -*-
import io
import json
import logging
from urllib.parse import quote
from odoo import http
from odoo.http import request
from datetime import datetime

_logger = logging.getLogger(__name__)

class CartDetectsController(http.Controller):
    """購物車檢測報告控制器"""

    # ==================== 類別常數 ====================

    # 基礎標題欄位
    BASE_COL_MAP = {
        "產品ID": "product_id",
        "產品名稱": "product_name",
        "姓名": "name",
    }

    # 人類醫學表單標題欄位
    HUMAN_COL_MAP = {
        "送檢單位": "inspection_unit",
        "性別(male:男、female:女、other:其他、unknown:不提供)": "gender",
        "病歷號碼": "medical_no",
        "出生年月日(2026-01-01)": "birth_date",
        "樣本類型(serum:血清、plasma:血漿)": "sample_type",
        "備註": "note",
    }

    # 小動物表單標題欄位
    ANIMAL_COL_MAP = {
        "送檢醫師": "doctor_name",
        "病患名稱": "patient_name",
        "物種(dog:狗、cat:貓)": "animal_type",
        "品系": "breed",
        "體重": "weight",
        "採血日期(2026-01-01)": "collect_date",
        "性別(male:男、female:女)": "gender",
        "是否絕育(1:是、0:否)": "neutered",
        "年齡": "age",
        "家中是否有飼養其他動物(1:是、0:否)": "has_other_animals",
        "其他動物數量(數字)": "other_animals_count",
        "跳蚤病史(infected:正在感染、treated:有但已除蚤、no:否、unknown:不確定)": "flea_history",
        "環境是否有改變(1:是、0:否)": "environment_changed",
        "簡述環境變化": "environment_change_desc",
        "初次發作年齡": "first_onset_age",
        "是否反覆發作(1:是、0:否)": "recurrent_years",
        "反覆發作持續時間": "recurrent_duration",
        "主要臨床類型(skin:皮膚型、respiratory:呼吸道、digestive:消化道)": "symptom_type",
        "一年至少發作三個月以上(1:是、0:否)": "long_term_attack",
        "皮膚症狀(可多選，請以逗號分隔)(搔癢、脫毛、皮膚紅疹、趾間炎、耳炎、結痂、抓癢、苔癬化、丘疹、膿皰、膿皮症、脂漏性皮膚炎、皮屑芽孢菌(馬拉色菌))": "skin_symptom_ids",
        "外寄生蟲種別": "parasite_type",
        "繼發性感染": "secondary_infection",
        "症狀好發期(seasonal:季節性、non_seasonal:整年都有非季節性、specific:特定季節更嚴重)": "symptom_period",
        "嚴重季節(spring:春季、summer:夏季、autumn:秋季、winter:冬季)": "severe_season",
        "或哪些月份特別嚴重": "severe_months",
        "Corticosteroids 藥物成分名、劑量和給予途徑": "corticosteroids_note",
        "Corticosteroids 最後一次給藥時間(2026-01-01)": "corticosteroids_last_date",
        "Corticosteroids 改善/無感/惡化(better:改善、same:無感、worse:惡化)": "corticosteroids_effect",
        "Antihistamine 藥物成分名、劑量和給予途徑": "antihistamine_note",
        "Antihistamine 最後一次給藥時間(2026-01-01)": "antihistamine_last_date",
        "Antihistamine 改善/無感/惡化(better:改善、same:無感、worse:惡化)": "antihistamine_effect",
        "Antibiotics 藥物成分名、劑量和給予途徑": "antibiotics_note",
        "Antibiotics 最後一次給藥時間(2026-01-01)": "antibiotics_last_date",
        "Antibiotics 改善/無感/惡化(better:改善、same:無感、worse:惡化)": "antibiotics_effect",
        "Antifungal agents 藥物成分名、劑量和給予途徑": "antifungal_note",
        "Antifungal agents 最後一次給藥時間(2026-01-01)": "antifungal_last_date",
        "Antifungal agents 改善/無感/惡化(better:改善、same:無感、worse:惡化)": "antifungal_effect",
        "Others 藥物成分名、劑量和給予途徑": "other_med_note",
        "Others 最後一次給藥時間(2026-01-01)": "other_med_last_date",
        "Others 改善/無感/惡化(better:改善、same:無感、worse:惡化)": "other_med_effect",
    }

    @classmethod
    def get_titles_by_category(cls, category: str) -> list:
        """根據類別取得完整的標題列表"""
        titles = list(cls.BASE_COL_MAP.keys())
        if category == "0":
            titles += list(cls.HUMAN_COL_MAP.keys())
        elif category == "1":
            titles += list(cls.ANIMAL_COL_MAP.keys())
        return titles

    # ==================== 私有輔助方法 ====================

    @staticmethod
    def _to_bool(value):
        """將字串轉換為布林值"""
        return value == "true" if value else False

    @staticmethod
    def _to_int(value, default=0):
        """將值轉換為整數"""
        return int(value) if value else default

    @staticmethod
    def _to_float(value, default=0.0):
        """將值轉換為浮點數"""
        return float(value) if value else default

    @staticmethod
    def _format_date(date_value):
        """格式化日期為字串"""
        return date_value.strftime("%Y-%m-%d") if date_value else ""

    @staticmethod
    def _build_human_data_dict(requisition):
        """組裝人類醫學送檢單資料字典"""
        return {
            "product_template_id": requisition.product_template_id.name or "",
            "internal_note": requisition.internal_note or "",
            "inspection_unit": requisition.inspection_unit or "",
            "patient_name": requisition.patient_name or "",
            "gender": requisition.gender or "",
            "medical_no": requisition.medical_no or "",
            "birth_date": CartDetectsController._format_date(requisition.birth_date),
            "sample_type": requisition.sample_type or "",
            "note": requisition.note or "",
        }

    @staticmethod
    def _build_animal_data_dict(requisition):
        """組裝小動物送檢單完整資料字典（用於返回前端）"""
        fmt = CartDetectsController._format_date
        return {
            "product_template_id": requisition.product_template_id.name or "",
            "internal_note": requisition.internal_note or "",
            "partner_id": requisition.partner_id.name or "",
            "doctor_name": requisition.doctor_name or "",
            "partner_address": requisition.partner_address or "",
            "partner_email": requisition.partner_email or "",
            "partner_phone": requisition.partner_phone or "",
            "patient_name": requisition.patient_name or "",
            "owner_name": requisition.owner_name or "",
            "animal_type": requisition.animal_type or "",
            "breed": requisition.breed or "",
            "weight": requisition.weight or 0,
            "collect_date": fmt(requisition.collect_date),
            "gender": requisition.gender or "",
            "neutered": requisition.neutered or False,
            "age": requisition.age or 0,
            "has_other_animals": requisition.has_other_animals or False,
            "other_animals_count": requisition.other_animals_count or "",
            "flea_history": requisition.flea_history or "",
            "environment_changed": requisition.environment_changed or False,
            "environment_change_desc": requisition.environment_change_desc or "",
            "first_onset_age": requisition.first_onset_age or 0,
            "recurrent_years": requisition.recurrent_years or False,
            "recurrent_duration": requisition.recurrent_duration or "",
            "symptom_type": requisition.symptom_type or "",
            "long_term_attack": requisition.long_term_attack or False,
            "skin_symptom_ids": [s.name for s in requisition.skin_symptom_ids],
            "parasite_type": requisition.parasite_type or "",
            "secondary_infection": requisition.secondary_infection or "",
            "symptom_period": requisition.symptom_period or "",
            "severe_season": requisition.severe_season or "",
            "severe_months": requisition.severe_months or "",
            "corticosteroids_note": requisition.corticosteroids_note or "",
            "corticosteroids_last_date": fmt(requisition.corticosteroids_last_date),
            "corticosteroids_effect": requisition.corticosteroids_effect or "",
            "antihistamine_note": requisition.antihistamine_note or "",
            "antihistamine_last_date": fmt(requisition.antihistamine_last_date),
            "antihistamine_effect": requisition.antihistamine_effect or "",
            "antibiotics_note": requisition.antibiotics_note or "",
            "antibiotics_last_date": fmt(requisition.antibiotics_last_date),
            "antibiotics_effect": requisition.antibiotics_effect or "",
            "antifungal_note": requisition.antifungal_note or "",
            "antifungal_last_date": fmt(requisition.antifungal_last_date),
            "antifungal_effect": requisition.antifungal_effect or "",
            "other_med_note": requisition.other_med_note or "",
            "other_med_last_date": fmt(requisition.other_med_last_date),
            "other_med_effect": requisition.other_med_effect or False,
        }

    @staticmethod
    def _build_human_requisition_vals(data):
        """從請求數據構建人類醫學送檢單欄位值"""
        return {
            "gender": data.get("gender") or False,
            "inspection_unit": data.get("inspection_unit"),
            "medical_no": data.get("medical_no"),
            "birth_date": data.get("birth_date") or False,
            "sample_type": data.get("sample_type"),
            "note": data.get("note"),
        }

    @staticmethod
    def _build_animal_requisition_vals(data):
        """從請求數據構建小動物送檢單欄位值"""
        to_bool = CartDetectsController._to_bool
        to_int = CartDetectsController._to_int
        to_float = CartDetectsController._to_float

        return {
            "doctor_name": data.get("doctor_name"),
            "patient_name": data.get("patient_name"),
            "animal_type": data.get("animal_type") or False,
            "breed": data.get("breed"),
            "weight": to_float(data.get("weight")),
            "collect_date": data.get("collect_date") or False,
            "gender": data.get("gender") or False,
            "neutered": to_bool(data.get("neutered")),
            "age": data.get("age"),
            "has_other_animals": to_bool(data.get("has_other_animals")),
            "other_animals_count": to_int(data.get("other_animals_count")),
            "flea_history": data.get("flea_history") or False,
            "environment_changed": to_bool(data.get("environment_changed")),
            "environment_change_desc": data.get("environment_change_desc"),
            "first_onset_age": data.get("first_onset_age"),
            "recurrent_years": to_bool(data.get("recurrent_years")),
            "recurrent_duration": data.get("recurrent_duration"),
            "symptom_type": data.get("symptom_type") or False,
            "long_term_attack": to_bool(data.get("long_term_attack")),
            "parasite_type": data.get("parasite_type"),
            "secondary_infection": data.get("secondary_infection"),
            "symptom_period": data.get("symptom_period") or False,
            "severe_season": data.get("severe_season") or False,
            "severe_months": data.get("severe_months"),
            "corticosteroids_note": data.get("corticosteroids_note"),
            "corticosteroids_last_date": data.get("corticosteroids_last_date") or False,
            "corticosteroids_effect": data.get("corticosteroids_effect") or False,
            "antihistamine_note": data.get("antihistamine_note"),
            "antihistamine_last_date": data.get("antihistamine_last_date") or False,
            "antihistamine_effect": data.get("antihistamine_effect") or False,
            "antibiotics_note": data.get("antibiotics_note"),
            "antibiotics_last_date": data.get("antibiotics_last_date") or False,
            "antibiotics_effect": data.get("antibiotics_effect") or False,
            "antifungal_note": data.get("antifungal_note"),
            "antifungal_last_date": data.get("antifungal_last_date") or False,
            "antifungal_effect": data.get("antifungal_effect") or False,
            "other_med_note": data.get("other_med_note"),
            "other_med_last_date": data.get("other_med_last_date") or False,
            "other_med_effect": data.get("other_med_effect") or False,
        }

    @staticmethod
    def _process_skin_symptom_ids(data):
        """處理 skin_symptom_ids Many2many 欄位，返回 ORM 命令"""
        skin_symptom_names = data.get("skin_symptom_ids", [])
        if not skin_symptom_names:
            return [(5, 0, 0)]

        if isinstance(skin_symptom_names, str):
            skin_symptom_names = [skin_symptom_names]

        symptom_ids = (
            request.env["idx.skin.symptom"]
            .sudo()
            .search([("name", "in", skin_symptom_names)])
            .ids
        )
        return [(6, 0, symptom_ids)] if symptom_ids else [(5, 0, 0)]

    # ==================== 路由方法 ====================

    @http.route(
        ["/shop/report/default_partner_name"],
        type="http",
        auth="user",
        methods=["GET"],
        website=True,
        csrf=False,
    )
    def get_default_partner_name(self):
        partner = request.env.user.partner_id
        commercial_partner = partner.commercial_partner_id or partner
        return request.make_json_response(
            {
                "success": True,
                "partner_name": commercial_partner.name or "",
            }
        )

    @http.route(
        ["/shop/reports/delete"],
        type="http",
        auth="user",
        methods=["DELETE"],
        website=True,
        csrf=False,
    )
    def delete_cart(self, order_id=None, **kw):
        """
        刪除購物車內所有檢測報告
        :param order_id: sale order ID
        :return: {success: True, deleted_count: int} or {error: str}
        """
        try:
            if not order_id:
                return request.make_json_response(
                    {"success": False, "error": "缺少 order_id 參數"}
                )

            reports = (
                request.env["sale.order.report"]
                .sudo()
                .search([("order_id", "=", int(order_id))])
            )

            reports.unlink()

            return request.make_json_response({"success": True})
        except Exception as e:
            return request.make_json_response({"success": False, "error": str(e)})

    @http.route(
        ["/shop/report/qty"],
        type="http",
        auth="user",
        methods=["GET"],
        website=True,
        csrf=False,
    )
    def get_report_qty(self, cart_id, cart_line_id, category=None, **kw):
        """
        取得購物車每筆單身的報告數量
        :param cart_id: sale order ID
        :param cart_line_id: sale order line ID
        :param category: report category
        :return: {success: True, report_qty: int} or {error: str, report_qty: None}
        """
        try:
            reports = (
                request.env["sale.order.report"]
                .sudo()
                .search(
                    [
                        ("order_id", "=", int(cart_id)),
                        ("order_line_id", "=", int(cart_line_id)),
                    ],
                    order="create_date desc",
                )
            )

            if not reports:
                return request.make_json_response(
                    {
                        "success": True,
                        "report_qty": 0,
                        "personDatas": {},
                        "animalDatas": {},
                    }
                )

            personDatas = {}
            animalDatas = {}
            for report in reports:
                if int(cart_line_id) not in personDatas:
                    personDatas[int(cart_line_id)] = []
                if int(cart_line_id) not in animalDatas:
                    animalDatas[int(cart_line_id)] = []

                # 人類醫學資料
                if category == "0":
                    requisitions = (
                        request.env["idx.test.requisition1"]
                        .sudo()
                        .search(
                            [("order_report_id", "=", report.id)],
                            order="create_date desc",
                        )
                    )
                    for requisition in requisitions:
                        data = self._build_human_data_dict(requisition)
                        data["report_id"] = report.id
                        personDatas[int(cart_line_id)].append(data)

                # 小動物資料
                elif category == "1":
                    requisitions = (
                        request.env["idx.test.requisition2"]
                        .sudo()
                        .search(
                            [("order_report_id", "=", report.id)],
                            order="create_date desc",
                        )
                    )
                    for requisition in requisitions:
                        data = self._build_animal_data_dict(requisition)
                        data["report_id"] = report.id
                        animalDatas[int(cart_line_id)].append(data)

            return request.make_json_response(
                {
                    "success": True,
                    "report_qty": len(reports),
                    "personDatas": personDatas,
                    "animalDatas": animalDatas,
                }
            )
        except Exception as e:
            return request.make_json_response(
                {
                    "error": str(e),
                    "report_qty": None,
                    "personDatas": {},
                    "animalDatas": {},
                }
            )

    @http.route(
        ["/shop/report/create"],
        type="http",
        auth="user",
        methods=["POST"],
        website=True,
        csrf=False,
    )
    def create_report(self, **kw):
        """
        創建檢測報告
        :return: {success: True, report_qty: int, report_id: int} or {error: str}
        """
        try:
            # 從請求體讀取 JSON 數據
            data = json.loads(request.httprequest.data.decode("utf-8"))

            cart_id = data.get("cart_id")
            cart_line_id = data.get("cart_line_id")
            category = data.get("category", None)

            if not (cart_id and cart_line_id) or category is None:
                return request.make_json_response({"error": "缺少必要參數"})

            # 根據類別處理不同的表單數據
            report_vals = {
                "order_id": int(cart_id),
                "order_line_id": int(cart_line_id),
            }
            requisition_vals = {
                "order_id": int(cart_id),
            }

            # 人類醫學表單欄位
            if category == "0":
                if not data.get("patient_name"):
                    return request.make_json_response(
                        {"error": "請確認姓名欄位是否填寫!"}
                    )
                report_vals["name"] = data.get("patient_name")
                requisition_vals.update(self._build_human_requisition_vals(data))

            # 小動物表單欄位
            elif category == "1":
                if not data.get("owner_name"):
                    return request.make_json_response(
                        {"error": "請確認飼主姓名欄位是否填寫!"}
                    )
                report_vals["name"] = data.get("owner_name")
                requisition_vals.update(self._build_animal_requisition_vals(data))
                requisition_vals["skin_symptom_ids"] = self._process_skin_symptom_ids(
                    data
                )

            try:
                # 創建報告記錄
                report = (
                    request.env["sale.order.report"]
                    .sudo()
                    .with_context(skip_ins_requisition=True)
                    .create(report_vals)
                )

                # 確認報告建立成功後才建立送檢單資料
                if not report or not report.id:
                    raise ValueError("報告建立失敗")

                requisition_vals.update({"order_report_id": report.id})

                # 根據 category 創建對應的送檢單
                if category == "0":
                    requisition = (
                        request.env["idx.test.requisition1"]
                        .sudo()
                        .create(requisition_vals)
                    )
                elif category == "1":
                    requisition = (
                        request.env["idx.test.requisition2"]
                        .sudo()
                        .create(requisition_vals)
                    )
                else:
                    raise ValueError(f"不支援的類別")

                if not requisition or not requisition.id:
                    raise ValueError("送檢單建立失敗")

            except Exception as inner_e:
                request.env.cr.rollback()
                return request.make_json_response(
                    {"error": f"資料寫入失敗: {str(inner_e)}"}
                )

            # 獲取更新後的報告數量
            report_count = (
                request.env["sale.order.report"]
                .sudo()
                .search_count([("order_line_id", "=", int(cart_line_id))])
            )

            # 組裝返回的送檢單數據
            if category == "0":
                requisition_data = self._build_human_data_dict(requisition)
            elif category == "1":
                requisition_data = self._build_animal_data_dict(requisition)
            else:
                requisition_data = {}

            return request.make_json_response(
                {
                    "success": True,
                    "report_id": report.id,
                    "report_qty": report_count,
                    "requisition": requisition_data,
                }
            )

        except Exception as e:
            return request.make_json_response({"error": str(e)})

    @http.route(
        ["/shop/report/delete"],
        type="http",
        auth="user",
        methods=["DELETE"],
        website=True,
        csrf=False,
    )
    def delete_report(self, report_id=None, **kw):
        """
        刪除檢測報告
        :param report_id: report ID
        :return: {success: True, report_qty: int} or {error: str}
        """
        try:
            if not report_id:
                return request.make_json_response({"error": "缺少 report_id 參數"})

            report = request.env["sale.order.report"].sudo().browse(int(report_id))

            if not report.exists():
                return request.make_json_response({"error": "找不到該報告"})

            # 保存 order_line_id 用於計算剩餘數量
            order_line_id = report.order_line_id.id

            # 刪除報告會連帶刪除關聯的送檢單（透過 ondelete='cascade'）
            report.unlink()

            # 計算刪除後的報告數量
            report_count = (
                request.env["sale.order.report"]
                .sudo()
                .search_count([("order_line_id", "=", order_line_id)])
            )

            return request.make_json_response(
                {
                    "success": True,
                    "report_qty": report_count,
                }
            )

        except Exception as e:
            return request.make_json_response({"error": str(e)})

    @http.route(
        ["/shop/report/update"],
        type="http",
        auth="user",
        methods=["PUT"],
        website=True,
        csrf=False,
    )
    def update_report(self, **kw):
        """
        更新檢測報告
        :return: {success: True, requisition: dict} or {error: str}
        """
        try:
            # 從請求體讀取 JSON 數據
            data = json.loads(request.httprequest.data.decode("utf-8"))

            report_id = data.get("report_id")
            category = data.get("category", None)

            if not report_id or category is None:
                return request.make_json_response({"error": "缺少必要參數"})

            # 查找報告
            report = request.env["sale.order.report"].sudo().browse(int(report_id))

            if not report.exists():
                return request.make_json_response({"error": "找不到該報告"})

            # 根據類別更新不同的表單數據
            if category == "0":
                if not data.get("patient_name"):
                    return request.make_json_response(
                        {"error": "請確認姓名欄位是否填寫!"}
                    )

                report.with_context(skip_ins_requisition=True).write(
                    {"name": data.get("patient_name")}
                )

                requisition = (
                    request.env["idx.test.requisition1"]
                    .sudo()
                    .search([("order_report_id", "=", report.id)], limit=1)
                )
                if not requisition:
                    return request.make_json_response({"error": "找不到對應的送檢單"})

                requisition.write(self._build_human_requisition_vals(data))
                return request.make_json_response(
                    {
                        "success": True,
                        "requisition": self._build_human_data_dict(requisition),
                    }
                )

            # 小動物更新邏輯
            elif category == "1":
                if not data.get("owner_name"):
                    return request.make_json_response(
                        {"error": "請確認飼主姓名欄位是否填寫!"}
                    )

                report.with_context(skip_ins_requisition=True).write(
                    {"name": data.get("owner_name")}
                )

                requisition = (
                    request.env["idx.test.requisition2"]
                    .sudo()
                    .search([("order_report_id", "=", report.id)], limit=1)
                )
                if not requisition:
                    return request.make_json_response({"error": "找不到對應的送檢單"})

                requisition_vals = self._build_animal_requisition_vals(data)
                requisition_vals["skin_symptom_ids"] = self._process_skin_symptom_ids(
                    data
                )
                requisition.write(requisition_vals)

                return request.make_json_response(
                    {
                        "success": True,
                        "requisition": self._build_animal_data_dict(requisition),
                    }
                )

            return request.make_json_response({"error": "不支援的類別"})

        except Exception as e:
            return request.make_json_response({"error": str(e)})

    @http.route(
        ["/shop/report/download_template"],
        type="http",
        auth="user",
        methods=["GET"],
        website=True,
        csrf=False,
    )
    def download_template(self, category=None, **kw):
        """
        下載 Excel 範本
        :param category: 類別
        :return: xlsx 檔案
        """
        try:
            import xlsxwriter
        except ImportError:
            return request.make_json_response({"error": "xlsxwriter 模組未安裝"})

        if not category:
            return request.make_json_response({"error": "缺少必要參數"})

        try:
            if category not in ("0", "1"):
                return request.make_json_response({"error": "不支援的類別"})

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {"in_memory": True})
            sheet_name = "人醫樣板" if category == "0" else "小動物樣板"
            worksheet = workbook.add_worksheet(sheet_name[:31])

            # 撈取使用者價格表中的有效商品，並依檢測單類別過濾
            pricelist = request.env.user.partner_id.property_product_pricelist
            if not pricelist:
                return request.make_json_response({"error": "找不到使用者價格表"})

            pricelist_items = (
                request.env["product.pricelist.item"]
                .sudo()
                .search(
                    [
                        ("pricelist_id", "=", pricelist.id),
                        ("product_tmpl_id", "!=", False),
                        "|",
                        ("date_start", "=", False),
                        ("date_start", "<=", datetime.now()),
                        "|",
                        ("date_end", "=", False),
                        ("date_end", ">=", datetime.now()),
                    ]
                )
            )

            products = request.env["product.template"].browse()
            for pricelist_item in pricelist_items:
                product = pricelist_item.product_tmpl_id
                if product and product.category == category:
                    products |= product

            if not products:
                category_label = "人醫" if category == "0" else "小動物"
                return request.make_json_response(
                    {"error": f"目前價格表中沒有可下載的{category_label}產品"}
                )

            titles = self.get_titles_by_category(category)

            if category == "0":
                sample_tail_rows = [
                    ["範例姓名1", "單位A", "male", "A123456789", "1990-01-01", "serum", "無"],
                    ["範例姓名2", "單位B", "other", "B987654321", "1985-12-31", "plasma", "無"],
                ]
            else:
                sample_tail_rows = [
                    [
                        "範例姓名1",
                        "範例醫師1",
                        "範例病患1",
                        "dog",
                        "拉布拉多",
                        "25.5",
                        "2026-01-01",
                        "male",
                        "1",
                        "5歲",
                        "1",
                        2,
                        "infected",
                        "1",
                        "搬家",
                        "2歲",
                        "1",
                        "半年",
                        "skin",
                        "1",
                        "皮屑芽孢菌(馬拉色菌)、搔癢、脫毛",
                        "跳蚤",
                        "無",
                        "seasonal",
                        "spring",
                        "3,4,5",
                        "Prednisone 5mg orally once daily",
                        "2026-01-01",
                        "better",
                        "Cetirizine 10mg orally once daily",
                        "2026-01-01",
                        "same",
                        "Amoxicillin 250mg orally twice daily",
                        "2026-01-01",
                        "worse",
                        "Ketoconazole 100mg orally once daily",
                        "2026-01-01",
                        "better",
                        "Vitamin E 100IU orally once daily",
                        "2026-01-01",
                        "same",
                    ],
                    [
                        "範例姓名2",
                        "範例醫師2",
                        "範例病患2",
                        "cat",
                        "波斯貓",
                        "4.2",
                        "2026-01-01",
                        "female",
                        "0",
                        "3歲",
                        "0",
                        0,
                        "no",
                        "0",
                        "無",
                        "1歲",
                        "0",
                        "無",
                        "respiratory",
                        "0",
                        "",
                        "耳疥蟲、結痂",
                        "無",
                        "non_seasonal",
                        "",
                        "",
                        "Dexamethasone 2mg orally once daily",
                        "2026-01-01",
                        "same",
                        "",
                        "",
                        "",
                        "Clindamycin 150mg orally twice daily",
                        "2026-01-01",
                        "better",
                        "",
                        "",
                        "",
                        "Omega-3 fatty acids 50mg orally once daily",
                        "2026-01-01",
                        "better",
                    ],
                ]

            for col_num, title in enumerate(titles):
                worksheet.write(0, col_num, title)

            for row_num, product in enumerate(products, start=1):
                sample_tail = sample_tail_rows[(row_num - 1) % len(sample_tail_rows)]
                row_values = [product.id, product.name or "", *sample_tail]
                for col_num, value in enumerate(row_values):
                    worksheet.write(row_num, col_num, value)

            workbook.close()
            output.seek(0)

            filename = (
                "送檢報告範本_人醫樣板.xlsx"
                if category == "0"
                else "送檢報告範本_小動物樣板.xlsx"
            )
            # 使用 RFC 5987 編碼處理中文檔名
            encoded_filename = quote(filename)
            return request.make_response(
                output.read(),
                headers=[
                    (
                        "Content-Type",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                    (
                        "Content-Disposition",
                        f"attachment; filename*=UTF-8''{encoded_filename}",
                    ),
                ],
            )
        except Exception as e:
            return request.make_json_response({"error": str(e)})

    @http.route(
        ["/shop/report/upload"],
        type="http",
        auth="user",
        methods=["POST"],
        website=True,
        csrf=False,
    )
    def upload_reports(self, **kw):
        """
        批次上傳檢測報告
        :return: {success: True, ...} or {error: str}
        """
        try:
            import openpyxl
        except ImportError:
            return request.make_json_response({"error": "openpyxl 模組未安裝"})

        # 取得上傳的檔案
        uploaded_file = kw.get("file")
        order_id = kw.get("order_id")

        if not uploaded_file:
            return request.make_json_response({"error": "未選擇檔案"})
        if not order_id or order_id == "undefined":
            sale_order = request.website.sale_get_order(force_create=True)
            if not sale_order:
                return request.make_json_response({"error": "無法建立購物車訂單"})
            order_id = sale_order.id

        # 驗證副檔名
        filename = uploaded_file.filename
        if not filename.lower().endswith(".xlsx"):
            return request.make_json_response(
                {"error": "檔案格式錯誤，僅支援 .xlsx 格式"}
            )

        try:
            # 讀取 Excel 檔案
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            worksheet = workbook.active

            # 讀取第一列標題
            header_row = [cell.value for cell in worksheet[1]]
            # 移除空值
            header_row = [h for h in header_row if h is not None]
            header_count = len(header_row)

            # 根據標題數量判斷類別並取得預期標題
            if header_count == len(CartDetectsController.BASE_COL_MAP) + len(
                CartDetectsController.HUMAN_COL_MAP
            ):
                category = "0"
                expected_titles = CartDetectsController.get_titles_by_category(category)
            elif header_count == len(CartDetectsController.BASE_COL_MAP) + len(
                CartDetectsController.ANIMAL_COL_MAP
            ):
                category = "1"
                expected_titles = CartDetectsController.get_titles_by_category(category)
            else:
                return request.make_json_response(
                    {"error": "檔案標題格式錯誤，請使用正確的範本檔案"}
                )

            # 上傳的 Excel 資料列
            data_rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for col_num in range(header_count):
                    header = header_row[col_num]
                    value = row[col_num] if col_num < len(row) else None
                    row_data[header] = value
                # 根據 expected_titles 的順序組裝欄位值
                ordered_row = {
                    CartDetectsController.BASE_COL_MAP.get(title)
                    or CartDetectsController.ANIMAL_COL_MAP.get(title)
                    or CartDetectsController.HUMAN_COL_MAP.get(title): row_data.get(
                        title, "-Err-"
                    )
                    for title in expected_titles
                }
                if "-Err-" in ordered_row:
                    return request.make_json_response(
                        {"error": "檔案標題名稱錯誤，請使用正確的範本檔案"}
                    )
                data_rows.append(ordered_row)
                
            pricelist = request.env.user.partner_id.property_product_pricelist
            if not pricelist:
                return request.make_json_response({"error": "找不到使用者價格表"})

            pricelist_items = (
                request.env["product.pricelist.item"]
                .sudo()
                .search(
                    [
                        ("pricelist_id", "=", pricelist.id),
                        ("product_tmpl_id", "!=", False),
                        "|",
                        ("date_start", "=", False),
                        ("date_start", "<=", datetime.now()),
                        "|",
                        ("date_end", "=", False),
                        ("date_end", ">=", datetime.now()),
                    ]
                )
            )

            products = []
            for pricelist_item in pricelist_items:
                product = pricelist_item.product_tmpl_id
                if product and product.category == category:
                    products.append(product.id)
            
            invalid_product_ids = []
            # 寫入送檢報告資料
            for data in data_rows:
                if not (
                    data.get("product_id")
                    and data.get("name")
                ):
                    raise ValueError("缺少必要的報告參數")
                

                if data.get("product_id") not in products:
                    _logger.info("產品 ID %s 不在使用者價格表中，已跳過", data.get("product_id"))
                    invalid_product_ids.append(str(data.get("product_id")))
                    continue  # 跳過不屬於此使用者的產品
                # 確認data.get("order_line_id")屬於此購物車訂單內
                sale_order_line = (
                    request.env["sale.order.line"]
                    .sudo()
                    .search(
                        [
                            ("product_template_id", "=", data.get("product_id")),
                            ("order_id", "=", int(order_id)),
                        ]
                    )
                )
                if not sale_order_line:
                    product_template = (
                        request.env["product.template"]
                        .sudo()
                        .browse(int(data.get("product_id")))
                        .exists()
                    )
                    if not product_template:
                        raise ValueError("找不到對應的產品，無法建立訂單明細")

                    sale_order_line = (
                        request.env["sale.order.line"]
                        .sudo()
                        .create(
                            {
                                "order_id": int(order_id),
                                "product_template_id": product_template.id,
                            }
                        )
                    )

                report = (
                    request.env["sale.order.report"]
                    .sudo()
                    .with_context(skip_ins_requisition=True)
                    .create(
                        {
                            "order_id": int(order_id),
                            "order_line_id": sale_order_line.id,
                            "name": data.get("name"),
                        }
                    )
                )

                # 確認報告建立成功後才建立送檢單資料
                if not report or not report.id:
                    raise ValueError("報告建立失敗")
                requisition_data = data.copy()
                del requisition_data["product_id"]
                del requisition_data["product_name"]
                del requisition_data["name"]

                # 人類醫學送檢單
                if report.category == "0":
                    requisition = (
                        request.env["idx.test.requisition1"]
                        .sudo()
                        .create(
                            {
                                **requisition_data,
                                "order_id": int(order_id),
                                "order_report_id": report.id,
                            }
                        )
                    )
                # 小動物送檢單
                elif report.category == "1":
                    skin_symptom_names = requisition_data.pop("skin_symptom_ids", [])
                    # 處理頓號分隔的症狀字串
                    if isinstance(skin_symptom_names, str):
                        skin_symptom_names = [
                            s.strip()
                            for s in skin_symptom_names.split("、")
                            if s.strip()
                        ]
                    skin_symptom_cmds = self._process_skin_symptom_ids(
                        {"skin_symptom_ids": skin_symptom_names}
                    )
                    requisition = (
                        request.env["idx.test.requisition2"]
                        .sudo()
                        .create(
                            {
                                **requisition_data,
                                "skin_symptom_ids": skin_symptom_cmds,
                                "order_id": int(order_id),
                                "order_report_id": report.id,
                            }
                        )
                    )

                if not requisition or not requisition.id:
                    raise ValueError("送檢單建立失敗")
            message = ""    
            if invalid_product_ids:
                invalid_product_ids = list(dict.fromkeys(invalid_product_ids))
                message = (
                    "報告批次上傳成功，但以下產品 ID 不在使用者價格表中，已略過："
                    + ", ".join(invalid_product_ids)
                )
            else:
                message = "報告批次上傳成功"

            return request.make_json_response(
                {
                    "success": True,
                    "message": message,
                    "warning_product_ids": invalid_product_ids,
                }
            )

        except Exception as e:
            request.env.cr.rollback()
            return request.make_json_response({"error": f"檔案批次上傳失敗: {str(e)}"})
